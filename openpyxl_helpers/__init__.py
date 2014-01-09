__author__ = 'sujeet'

import openpyxl
import numpy
from openpyxl.style import NumberFormat
from functools import partial

def read_table_from_named_range(workbook,name):
	nr = workbook.get_named_range(name)
	if nr is None:
		raise Exception("Error: no such named range")
	else:
		ws = nr.destinations[0][0]
		cellcol = ws.range(nr.destinations[0][1])
		outputTable = []
		for row in cellcol[1:]:
			outputTable.append(dict(zip([c.value.lower() for c in cellcol[0]],[c.value for c in row])))
		return outputTable

def read_value_from_named_range(workbook,name):
	nr = workbook.get_named_range(name)
	if nr is None:
		raise Exception("Error: no such named range")
	else:
		ws = nr.destinations[0][0]
		cell = ws.range(nr.destinations[0][1])
		if type(cell) is not openpyxl.cell.Cell:
			raise Exception("Function 'read_value_from_named_range' can only be used for single cell named ranges")
		else:
			return cell.value

def write_list_of_values(worksheet, starting_range, list, header=None, number_format=NumberFormat.FORMAT_GENERAL):
	if header is not None:
		worksheet.cell(coordinate=starting_range).value = header
		i=1
	else:
		i=0

	for val in list:
		worksheet.cell(coordinate=starting_range).offset(i,0).value = val
		worksheet.cell(coordinate=starting_range).offset(i,0).style.number_format.format_code = number_format
		i = i + 1

def create_distribution_table(worksheet, starting_range, value,number_of_bins=10,number_format=NumberFormat.FORMAT_GENERAL):
	number_format=number_format.replace("\"","")
	bins = numpy.arange(0,numpy.mean(value)*2,numpy.mean(value)*2/number_of_bins)
	bins = numpy.append(bins,[99999999])

	hist = numpy.histogram(value,bins)
	dispBins = []
	for i in range(0,len(bins)-2):
		rangeVal = '=Text(%(start)s,"%(format)s")&" - "&Text(%(end)s,"%(format)s")' % dict(start=bins[i],
		                                                                                   end=bins[i+1],format=number_format)
		dispBins.append(rangeVal)
	dispBins.append('=Text(%(start)s,"%(format)s")&" +"' % dict(start=bins[len(bins)-2],format=number_format))

	worksheet.cell(starting_range).value = "Distribution Table"
	next_range = worksheet.cell(starting_range).offset(1,0).address
	write_list_of_values(worksheet, next_range, dispBins,'Bins',NumberFormat.FORMAT_GENERAL)
	write_list_of_values(worksheet, worksheet.cell(next_range).offset(0,1).address, list(hist[0]),'# of Reps',NumberFormat.FORMAT_GENERAL)

def create_stats_table(worksheet, starting_range, value,number_format=NumberFormat.FORMAT_GENERAL):
	worksheet.cell(starting_range).value = "Stats Table"
	next_range = worksheet.cell(starting_range).offset(1,0).address

	functions = [
		{'label':'Sum','func':numpy.sum},
		{'label':'Mean','func':numpy.mean},
		{'label':'Standard Deviation','func':numpy.std},
		{'label':'Median','func':numpy.median},
		{'label':'Max','func':numpy.max},
		{'label':'Min','func':numpy.min},
		{'label':'10th Percentile','func':partial(numpy.percentile,q=10)},
		{'label':'25th Percentile','func':partial(numpy.percentile,q=25)},
		{'label':'50th Percentile','func':partial(numpy.percentile,q=50)},
		{'label':'75th Percentile','func':partial(numpy.percentile,q=75)},
		{'label':'90th Percentile','func':partial(numpy.percentile,q=90)},
		{'label':'90th over 10th Percentile','func':lambda x: numpy.percentile(x,q=90)/numpy.percentile(x,q=10),'format':NumberFormat.FORMAT_NUMBER_00},
		{'label':'What % of total payout does top 10% reps take home','func':lambda x: numpy.sum(x[x>numpy.percentile(x,q=90)])/numpy.sum(x),'format': NumberFormat.FORMAT_PERCENTAGE_00},
	]

	i=0
	for f in functions:
		worksheet.cell(next_range).offset(i,0).value = f['label']
		worksheet.cell(next_range).offset(i,0).style.alignment.horizontal = 'right'
		worksheet.cell(next_range).offset(i,1).value = f['func'](value)
		worksheet.cell(next_range).offset(i,1).style.number_format.format_code= f['format'] if 'format' in f else number_format
		i = i + 1
